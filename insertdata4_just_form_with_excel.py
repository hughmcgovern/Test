# -*- coding: utf-8 -*-
"""
Created on Fri Oct 14 11:51:15 2022
babagoosh
@author: hughm
"""
from openpyxl import *
from tkinter import *

import ctypes  # An included library with Python install.

 # opening the existing excel file
wb = load_workbook(r"C:\Users\hughm\Documents\Hughs Documents\03. Day Jobs\Dorset College\Programming Essentials 2\excel.xlsx")

# create the sheet object
sheet = wb.active

def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30

    # write given insertdata to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Address"
    sheet.cell(row=1, column=3).value = "Phone"
    sheet.cell(row=1, column=4).value = "Email"
excel()   
def insert(name,address,phone,email):
    
    # assigning the max row and max column
    # value upto which insertdata is written
    # in an excel sheet to the variable
    current_row = sheet.max_row
    current_column = sheet.max_column

    # get method returns current text
    # as string which we write into
    # excel spreadsheet at particular location
    sheet.cell(row=current_row + 1, column=1).value = e1.get()
    sheet.cell(row=current_row + 1, column=2).value = e2.get()
    sheet.cell(row=current_row + 1, column=3).value = e3.get()
    sheet.cell(row=current_row + 1, column=4).value = e4.get()

    # save the file
    wb.save(r'C:\Users\hughm\Documents\Hughs Documents\03. Day Jobs\Dorset College\Programming Essentials 2\excel.xlsx')
    
    # close the file
    wb.close()
    
    
def Mbox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)

def Popup_InsertExcel(name,address,telephone,email):
    Mbox('Details', e1.get()+"\n"+e2.get()+"\n"+e3.get()+"\n"+e4.get()+"\n" + 'inserted into excel.xlsx', 1)
    insert(e1.get(),e2.get(),e3.get(),e4.get())
    clear()

def clear():
    e1.delete(0,END)
    e2.delete(0,END)
    e3.delete(0,END)
    e4.delete(0,END)

# this is form piece
master = Tk()

master.title('Record Insert')
master.configure(bg='cyan')

Label(master, bg='cyan', text="Name",padx=30, pady=30, font=("Garamond", 15)).grid(row=2)
Label(master, bg='cyan', text="Address",padx=30, pady=30, font=("Garamond", 15)).grid(row=3)
Label(master, bg='cyan', text="Telephone",padx=30, pady=30, font=("Garamond", 15)).grid(row=4)
Label(master, bg='cyan', text="Email",padx=30, pady=30, font=("Garamond", 15)).grid(row=5)

e1 = Entry(master,width=20, font=("Garamond", 15)) # this are the input boxes
e2 = Entry(master,width=20, font=("Garamond", 15))
e3 = Entry(master,width=20, font=("Garamond", 15))
e4 = Entry(master,width=20, font=("Garamond", 15))

e1.grid(row=2, column=5)
e2.grid(row=3, column=5)
e3.grid(row=4, column=5)
e4.grid(row=5, column=5)

Button(master,font=("Garamond", 15), text='Save', \
       command=lambda: Popup_InsertExcel(e1.get(),e2.get(),e3.get(),e4.get())).grid(row=7, column=6, sticky=W, padx=10,pady=5)

# x and y are the coordinates of the upper left corner
w = 450
h = 400
x = 0
y = 0
# use width x height + x_offset + y_offset (no spaces!)
master.geometry("%dx%d+%d+%d" % (w, h, x, y))

windowWidth = master.winfo_reqwidth()
windowHeight = master.winfo_reqheight()

# Gets both half the screen width/height and window width/height
positionRight = int(master.winfo_screenwidth()/2 - windowWidth/2)
positionDown = int(master.winfo_screenheight()/2 - windowHeight/2)

# Positions the window in the center of the page.
master.geometry("+{}+{}".format(positionRight-150, positionDown-150))


master.mainloop()
